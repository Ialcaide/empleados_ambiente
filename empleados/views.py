from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Cargo, Empleado
from .forms import CargoForm, EmpleadoForm
from django.core.paginator import Paginator
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse

@login_required
def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Empleados'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
    header_alignment = Alignment(horizontal='center')

    # Encabezados
    headers = ['#', 'Nombres', 'Apellidos', 'Correo', 'Sueldo', 'Fecha Ingreso', 'Cargo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Datos
    empleados = Empleado.objects.select_related('cargo').all()
    for row, emp in enumerate(empleados, 2):
        ws.cell(row=row, column=1, value=emp.id)
        ws.cell(row=row, column=2, value=emp.nombres)
        ws.cell(row=row, column=3, value=emp.apellidos)
        ws.cell(row=row, column=4, value=emp.correo)
        ws.cell(row=row, column=5, value=float(emp.sueldo))
        ws.cell(row=row, column=6, value=str(emp.fecha_ingreso))
        ws.cell(row=row, column=7, value=str(emp.cargo))

    # Ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="empleados.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="empleados.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph('Lista de Empleados', styles['Title']))

    # Datos
    data = [['#', 'Nombres', 'Apellidos', 'Correo', 'Sueldo', 'Fecha Ingreso', 'Cargo']]
    empleados = Empleado.objects.select_related('cargo').all()
    for emp in empleados:
        data.append([
            str(emp.id),
            emp.nombres,
            emp.apellidos,
            emp.correo,
            f'${emp.sueldo}',
            str(emp.fecha_ingreso),
            str(emp.cargo),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D1B2A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0dce8')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def exportar_cargos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cargos'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
    header_alignment = Alignment(horizontal='center')

    headers = ['#', 'Nombre', 'Descripción', 'Total Empleados']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    cargos = Cargo.objects.all()
    for row, cargo in enumerate(cargos, 2):
        ws.cell(row=row, column=1, value=cargo.id)
        ws.cell(row=row, column=2, value=cargo.nombre)
        ws.cell(row=row, column=3, value=cargo.descripcion)
        ws.cell(row=row, column=4, value=cargo.empleado_set.count())

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cargos.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_cargos_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cargos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Lista de Cargos', styles['Title']))

    data = [['#', 'Nombre', 'Descripción', 'Total Empleados']]
    cargos = Cargo.objects.all()
    for cargo in cargos:
        data.append([
            str(cargo.id),
            cargo.nombre,
            cargo.descripcion or '',
            str(cargo.empleado_set.count()),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D1B2A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0dce8')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

# ─── CARGOS ───────────────────────────────────────────

@login_required
def cargo_lista(request):
    cargos = Cargo.objects.all()
    query = request.GET.get('q', '')
    if query:
        cargos = cargos.filter(
            models.Q(nombre__icontains=query) |
            models.Q(descripcion__icontains=query)
        )

    paginator = Paginator(cargos, 10)  # 10 cargos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'empleados/cargo_lista.html', {
        'cargos': page_obj,
        'page_obj': page_obj,
        'query': query,
    })

@login_required
def cargo_crear(request):
    form = CargoForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Cargo creado exitosamente.')
        return redirect('cargo_lista')
    return render(request, 'empleados/cargo_form.html', {'form': form, 'titulo': 'Nuevo Cargo'})

@login_required
def cargo_editar(request, pk):
    cargo = get_object_or_404(Cargo, pk=pk)
    form = CargoForm(request.POST or None, instance=cargo)
    if form.is_valid():
        form.save()
        messages.success(request, f'Cargo "{cargo.nombre}" actualizado exitosamente.')
        return redirect('cargo_lista')
    return render(request, 'empleados/cargo_form.html', {'form': form, 'titulo': 'Editar Cargo'})

@login_required
def cargo_eliminar(request, pk):
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        nombre = cargo.nombre
        cargo.delete()
        messages.success(request, f'Cargo "{nombre}" eliminado exitosamente.')
        return redirect('cargo_lista')
    return render(request, 'empleados/cargo_confirmar_eliminar.html', {'objeto': cargo})

# ─── EMPLEADOS ────────────────────────────────────────

@login_required
def empleado_lista(request):
    empleados = Empleado.objects.select_related('cargo').all()
    
    query = request.GET.get('q', '')
    cargo_filtro = request.GET.get('cargo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    sueldo_min = request.GET.get('sueldo_min', '')
    sueldo_max = request.GET.get('sueldo_max', '')
    estado = request.GET.get('estado', '')

    if query:
        empleados = empleados.filter(
            models.Q(nombres__icontains=query) |
            models.Q(apellidos__icontains=query)
        )
    if cargo_filtro:
        empleados = empleados.filter(cargo__id=cargo_filtro)
    if fecha_desde:
        empleados = empleados.filter(fecha_ingreso__gte=fecha_desde)
    if fecha_hasta:
        empleados = empleados.filter(fecha_ingreso__lte=fecha_hasta)
    if sueldo_min:
        empleados = empleados.filter(sueldo__gte=sueldo_min)
    if sueldo_max:
        empleados = empleados.filter(sueldo__lte=sueldo_max)
    if estado == 'activo':
        empleados = empleados.filter(activo=True)
    elif estado == 'inactivo':
        empleados = empleados.filter(activo=False)

    paginator = Paginator(empleados, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    cargos = Cargo.objects.all()
    return render(request, 'empleados/empleado_lista.html', {
        'empleados': page_obj,
        'page_obj': page_obj,
        'cargos': cargos,
        'query': query,
        'cargo_filtro': cargo_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'sueldo_min': sueldo_min,
        'sueldo_max': sueldo_max,
        'estado': estado,
    })

@login_required
def empleado_crear(request):
    form = EmpleadoForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Empleado creado exitosamente.')
        return redirect('empleado_lista')
    return render(request, 'empleados/empleado_form.html', {'form': form, 'titulo': 'Nuevo Empleado'})

@login_required
def empleado_editar(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    form = EmpleadoForm(request.POST or None, instance=empleado)
    if form.is_valid():
        form.save()
        messages.success(request, f'Empleado "{empleado.nombres}" actualizado exitosamente.')
        return redirect('empleado_lista')
    return render(request, 'empleados/empleado_form.html', {'form': form, 'titulo': 'Editar Empleado'})

@login_required
def empleado_eliminar(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == 'POST':
        nombre = f"{empleado.nombres} {empleado.apellidos}"
        empleado.delete()
        messages.success(request, f'Empleado "{nombre}" eliminado exitosamente.')
        return redirect('empleado_lista')
    return render(request, 'empleados/empleado_confirmar_eliminar.html', {'objeto': empleado})

@login_required
def empleado_detalle(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    return render(request, 'empleados/empleado_detalle.html', {'empleado': empleado})

@login_required
def cargo_detalle(request, pk):
    cargo = get_object_or_404(Cargo, pk=pk)
    empleados = Empleado.objects.filter(cargo=cargo)
    return render(request, 'empleados/cargo_detalle.html', {'cargo': cargo, 'empleados': empleados})
