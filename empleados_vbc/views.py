from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.views import View
from empleados.models import Cargo, Empleado
from empleados.forms import CargoForm, EmpleadoForm
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

# ─── DETALLE ──────────────────────────────────────────

class EmpleadoDetailView(LoginRequiredMixin, DetailView):
    model = Empleado
    template_name = 'empleados_vbc/empleado_detalle.html'
    context_object_name = 'empleado'

class CargoDetailView(LoginRequiredMixin, DetailView):
    model = Cargo
    template_name = 'empleados_vbc/cargo_detalle.html'
    context_object_name = 'cargo'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['empleados'] = Empleado.objects.filter(cargo=self.object)
        return context

# ─── CARGOS ───────────────────────────────────────────

class CargoListView(LoginRequiredMixin, ListView):
    model = Cargo
    template_name = 'empleados_vbc/cargo_lista.html'
    context_object_name = 'cargos'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q', '')
        if query:
            queryset = queryset.filter(nombre__icontains=query) | queryset.filter(descripcion__icontains=query)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        return context

class CargoCreateView(LoginRequiredMixin, CreateView):
    model = Cargo
    form_class = CargoForm
    template_name = 'empleados_vbc/cargo_form.html'
    success_url = reverse_lazy('vbc_cargo_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Cargo (VBC)'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Cargo creado exitosamente.')
        return super().form_valid(form)

class CargoUpdateView(LoginRequiredMixin, UpdateView):
    model = Cargo
    form_class = CargoForm
    template_name = 'empleados_vbc/cargo_form.html'
    success_url = reverse_lazy('vbc_cargo_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Cargo (VBC)'
        return context

    def form_valid(self, form):
        messages.success(self.request, f'Cargo "{self.object.nombre}" actualizado exitosamente.')
        return super().form_valid(form)

class CargoDeleteView(LoginRequiredMixin, DeleteView):
    model = Cargo
    template_name = 'empleados_vbc/cargo_confirmar_eliminar.html'
    success_url = reverse_lazy('vbc_cargo_lista')
    context_object_name = 'objeto'

    def form_valid(self, form):
        messages.success(self.request, f'Cargo "{self.object.nombre}" eliminado exitosamente.')
        return super().form_valid(form)

# ─── EMPLEADOS ────────────────────────────────────────

class EmpleadoListView(LoginRequiredMixin, ListView):
    model = Empleado
    template_name = 'empleados_vbc/empleado_lista.html'
    context_object_name = 'empleados'
    paginate_by = 10

    def get_queryset(self):
        queryset = Empleado.objects.select_related('cargo').all()
        query = self.request.GET.get('q', '')
        cargo_filtro = self.request.GET.get('cargo', '')
        fecha_desde = self.request.GET.get('fecha_desde', '')
        fecha_hasta = self.request.GET.get('fecha_hasta', '')
        sueldo_min = self.request.GET.get('sueldo_min', '')
        sueldo_max = self.request.GET.get('sueldo_max', '')
        estado = self.request.GET.get('estado', '')
        if query:
            queryset = queryset.filter(nombres__icontains=query) | queryset.filter(apellidos__icontains=query)
        if cargo_filtro:
            queryset = queryset.filter(cargo__id=cargo_filtro)
        if fecha_desde:
            queryset = queryset.filter(fecha_ingreso__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_ingreso__lte=fecha_hasta)
        if sueldo_min:
            queryset = queryset.filter(sueldo__gte=sueldo_min)
        if sueldo_max:
            queryset = queryset.filter(sueldo__lte=sueldo_max)
        if estado == 'activo':
            queryset = queryset.filter(activo=True)
        elif estado == 'inactivo':
            queryset = queryset.filter(activo=False)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['cargo_filtro'] = self.request.GET.get('cargo', '')
        context['fecha_desde'] = self.request.GET.get('fecha_desde', '')
        context['fecha_hasta'] = self.request.GET.get('fecha_hasta', '')
        context['sueldo_min'] = self.request.GET.get('sueldo_min', '')
        context['sueldo_max'] = self.request.GET.get('sueldo_max', '')
        context['estado'] = self.request.GET.get('estado', '')
        context['cargos'] = Cargo.objects.all()
        return context

class EmpleadoCreateView(LoginRequiredMixin, CreateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'empleados_vbc/empleado_form.html'
    success_url = reverse_lazy('vbc_empleado_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Empleado (VBC)'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Empleado creado exitosamente.')
        return super().form_valid(form)

class EmpleadoUpdateView(LoginRequiredMixin, UpdateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'empleados_vbc/empleado_form.html'
    success_url = reverse_lazy('vbc_empleado_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Empleado (VBC)'
        return context

    def form_valid(self, form):
        messages.success(self.request, f'Empleado "{self.object.nombres}" actualizado exitosamente.')
        return super().form_valid(form)

class EmpleadoDeleteView(LoginRequiredMixin, DeleteView):
    model = Empleado
    template_name = 'empleados_vbc/empleado_confirmar_eliminar.html'
    success_url = reverse_lazy('vbc_empleado_lista')
    context_object_name = 'objeto'

    def form_valid(self, form):
        messages.success(self.request, f'Empleado "{self.object.nombres}" eliminado exitosamente.')
        return super().form_valid(form)

# ─── EXPORTAR ─────────────────────────────────────────

class ExportarExcelView(LoginRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Empleados'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0D2A1B', end_color='0D2A1B', fill_type='solid')
        header_alignment = Alignment(horizontal='center')

        headers = ['#', 'Nombres', 'Apellidos', 'Correo', 'Sueldo', 'Fecha Ingreso', 'Cargo', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        empleados = Empleado.objects.select_related('cargo').all()
        for row, emp in enumerate(empleados, 2):
            ws.cell(row=row, column=1, value=emp.id)
            ws.cell(row=row, column=2, value=emp.nombres)
            ws.cell(row=row, column=3, value=emp.apellidos)
            ws.cell(row=row, column=4, value=emp.correo)
            ws.cell(row=row, column=5, value=float(emp.sueldo))
            ws.cell(row=row, column=6, value=str(emp.fecha_ingreso))
            ws.cell(row=row, column=7, value=str(emp.cargo))
            ws.cell(row=row, column=8, value='Activo' if emp.activo else 'Inactivo')

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="empleados_vbc.xlsx"'
        wb.save(response)
        return response

class ExportarPdfView(LoginRequiredMixin, View):
    def get(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="empleados_vbc.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('Lista de Empleados (VBC)', styles['Title']))

        data = [['#', 'Nombres', 'Apellidos', 'Correo', 'Sueldo', 'Fecha Ingreso', 'Cargo', 'Estado']]
        empleados = Empleado.objects.select_related('cargo').all()
        for emp in empleados:
            data.append([
                str(emp.id),
                emp.nombres,
                emp.apellidos,
                emp.correo,
                f'${emp.sueldo}',
                str(emp.fecha_ingreso),
                str(emp.cargo),
                'Activo' if emp.activo else 'Inactivo',
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D2A1B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0dce8')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        doc.build(elements)
        return response
    
class ExportarCargosExcelView(LoginRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cargos'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0D2A1B', end_color='0D2A1B', fill_type='solid')
        header_alignment = Alignment(horizontal='center')

        headers = ['#', 'Nombre', 'Descripción', 'Total Empleados']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        cargos = Cargo.objects.all()
        for row, cargo in enumerate(cargos, 2):
            ws.cell(row=row, column=1, value=cargo.id)
            ws.cell(row=row, column=2, value=cargo.nombre)
            ws.cell(row=row, column=3, value=cargo.descripcion)
            ws.cell(row=row, column=4, value=cargo.empleado_set.count())

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="cargos_vbc.xlsx"'
        wb.save(response)
        return response

class ExportarCargosPdfView(LoginRequiredMixin, View):
    def get(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="cargos_vbc.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('Lista de Cargos (VBC)', styles['Title']))

        data = [['#', 'Nombre', 'Descripción', 'Total Empleados']]
        cargos = Cargo.objects.all()
        for cargo in cargos:
            data.append([
                str(cargo.id),
                cargo.nombre,
                cargo.descripcion or '',
                str(cargo.empleado_set.count()),
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D2A1B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0dce8')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        doc.build(elements)
        return response